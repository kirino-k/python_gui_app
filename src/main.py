#########################
##### Version 0.5.0 #####
#########################

import os
import re
import datetime
import openpyxl
import numpy as np
import random
import string
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl.styles import PatternFill


# ファイル選択ダイアログに最初に表示する path を設定する
input_book_dir = os.path.abspath(os.path.dirname(__file__))
ledger_book_dir = os.path.join(input_book_dir, "..")


def select_ledger_book():
    file_path = filedialog.askopenfilename(
        initialdir=ledger_book_dir, filetypes=[("Excel Files", ".xlsx")]
    )
    if file_path:
        ledger_book_name.set(file_path)


def select_input_book():
    file_path = filedialog.askopenfilename(
        initialdir=input_book_dir, filetypes=[("Excel Files", ".xlsx")]
    )
    if file_path:
        input_book_name.set(file_path)


def select_user_list():
    file_path = filedialog.askopenfilename(
        initialdir=input_book_dir, filetypes=[("Excel Files", ".xlsx")]
    )
    if file_path:
        user_list_name.set(file_path)


def run_process():

    return 0

    # # 入力値エラーをカウントする変数の初期化
    # fill_error = 0
    # value_error = 0

    # # メッセージの初期化
    # message = ""

    # # try:
    # # 台帳bookと申請フォームbookの読み込み
    # ledger_book = openpyxl.load_workbook(ledger_book_name.get())
    # input_book = openpyxl.load_workbook(input_book_name.get())

    # # 台帳bookの規格チェック
    # if (
    #     np.count_nonzero(
    #         list(
    #             map(
    #                 lambda x: x == "プロジェクト" or x == "ユーザーアカウント",
    #                 ledger_book.sheetnames,
    #             )
    #         )
    #     )
    #     != 2
    # ):

    #     # エラーメッセージ表示、アプリは終了しない
    #     messagebox.showerror("エラー", "台帳ファイルのフォーマットが規格と異なります")
    #     return 1

    # # 申請フォームbookの規格チェック
    # if (
    #     np.count_nonzero(
    #         list(
    #             map(
    #                 lambda x: x == "お読みください" or x == "基本事項の入力" or x == "ユーザーアカウントの入力",
    #                 input_book.sheetnames,
    #             )
    #         )
    #     )
    #     != 3
    # ):

    #     # エラーメッセージ表示、アプリは終了しない
    #     messagebox.showerror("エラー", "申請フォームのフォーマットが規格と異なります")
    #     return 1

    # # プロジェクト基本情報の入力を取得
    # project_input = input_book["基本事項の入力"]
    # project_id = project_input.cell(row=3, column=1).value
    # project_name = project_input.cell(row=3, column=2).value

    # # REDCap研究番号が未記載
    # if project_id is None:
    #     fill_error += 1
    #     message += "REDCap研究番号が未記載です。\n"
    #     project_input.cell(row=3, column=1).fill = PatternFill(
    #         patternType="solid", fgColor="FFFF00"
    #     )

    # # REDCap研究番号の記載がREDCap_数字4桁またはREDCap数字4桁ではない
    # elif (
    #     project_id is not None
    #     and re.match(r"^(RED)_?[0-9]{4}$", project_id.strip()) is None
    # ):
    #     value_error += 1
    #     message += "REDCap研究番号を正しく入力してください。\n"
    #     project_input.cell(row=3, column=1).fill = PatternFill(
    #         patternType="solid", fgColor="FFFF00"
    #     )

    # # ユーザーアカウント台帳の内容を取得
    # account_ledger = ledger_book["ユーザーアカウント"]
    # ledger_values = []
    # for row in account_ledger.iter_rows(min_row=2):
    #     ledger_value = [col.value for col in row]
    #     if ledger_value[0] is None:
    #         break
    #     ledger_values.append(ledger_value)

    # # ledger_value
    # #  0: 処理日時
    # #  1: REDCap研究番号
    # #  2: 確定ユーザ名
    # #  3: First Name
    # #  4: Last Name
    # #  5: メールアドレス
    # #  6: 氏名 (漢字)
    # #  7: 所属施設
    # #  8: 所属部署
    # #  9: トレーニング受講日
    # # 10: プロジェクト作成・コピー作成の申請権限

    # # ユーザーアカウント情報の入力を取得
    # input_values = []
    # account_input = input_book["ユーザーアカウントの入力"]
    # for row in account_input.iter_rows(min_row=3):
    #     input_value = [col.value for col in row]

    #     # ２セル以上の記入がある列は情報読み込み, 2セル未満の書き込みの場合はループを出る
    #     if np.count_nonzero(list(map(lambda x: x is not None, input_value))) < 2:
    #         break
    #     else:
    #         input_values.append(input_value)

    # # input_value
    # #  0: 成育REDCapアカウント
    # #  1: 希望ユーザ名 または 使用中ユーザ名
    # #  2: 氏
    # #  3: 名
    # #  4: 氏
    # #  5: 名
    # #  6: メールアドレス
    # #  7: 所属施設①
    # #  8: 所属施設②
    # #  9: 所属部署
    # # 10: トレーニング受講日
    # # 11: 連絡担当者
    # # 12: プロジェクト・コピー作成の申請権限の付与希望

    # # row-wiseな入力情報の精査
    # for index, input_value in enumerate(input_values):

    #     # 成育REDCapアカウント所持の有無が空欄
    #     if input_value[0] is None:
    #         fill_error += 1
    #         message += "行番号 " + str(index + 3) + ": 「成育REDCapアカウント」が未記載です。\n"
    #         account_input.cell(row=index + 3, column=1).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # 希望ユーザ名または使用中のユーザ名が空欄
    #     if input_value[1] is None:
    #         fill_error += 1
    #         message += "行番号 " + str(index + 3) + ": 「希望ユーザ名または使用中のユーザ名」が未記載です。\n"
    #         account_input.cell(row=index + 3, column=2).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # 希望ユーザ名または使用中のユーザ名に半角英数字, - (ハイフン), _ (アンダースコア) 以外の文字が含まれる
    #     elif re.match(r"^[a-zA-Z0-9_-]+$", input_value[1].strip()) is None:
    #         value_error += 1
    #         message += (
    #             "行番号 " + str(index + 3) + ": 「希望ユーザ名または使用中のユーザ名」に使用できない文字が含まれています。\n"
    #         )
    #         account_input.cell(row=index + 3, column=2).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # 希望ユーザ名が10文字以内でない
    #     elif input_value[0] == "持っていない" and len(input_value[1].strip()) > 10:
    #         value_error += 1
    #         message += (
    #             "行番号 " + str(index + 3) + ": 「希望ユーザ名または使用中のユーザ名」が既定の文字数を超えています。\n"
    #         )
    #         account_input.cell(row=index + 3, column=2).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # 使用中のユーザ名が12文字以内でない
    #     elif input_value[0] == "持っている" and len(input_value[1].strip()) > 12:
    #         value_error += 1
    #         message += (
    #             "行番号 " + str(index + 3) + ": 「希望ユーザ名または使用中のユーザ名」が既定の文字数を超えています。\n"
    #         )
    #         account_input.cell(row=index + 3, column=2).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # 氏 (日本語) が空欄
    #     if input_value[2] is None:
    #         fill_error += 1
    #         message += "行番号 " + str(index + 3) + ": 「氏 (日本語)」が未記載です。\n"
    #         account_input.cell(row=index + 3, column=3).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # 名 (日本語) が空欄
    #     if input_value[3] is None:
    #         fill_error += 1
    #         message += "行番号 " + str(index + 3) + ": 「名 (日本語)」が未記載です。\n"
    #         account_input.cell(row=index + 3, column=4).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # 氏 (ローマ字) が空欄
    #     if input_value[4] is None:
    #         fill_error += 1
    #         message += "行番号 " + str(index + 3) + ": 「氏 (ローマ字)」が未記載です。\n"
    #         account_input.cell(row=index + 3, column=5).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # 名 (ローマ字) が空欄
    #     if input_value[5] is None:
    #         fill_error += 1
    #         message += "行番号 " + str(index + 3) + ": 「名 (ローマ字)」が未記載です。\n"
    #         account_input.cell(row=index + 3, column=6).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # メールアドレスが空欄
    #     if input_value[6] is None:
    #         fill_error += 1
    #         message += "行番号 " + str(index + 3) + ": 「メールアドレス」が未記載です。\n"
    #         account_input.cell(row=index + 3, column=7).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # メールアドレスがメールアドレスの正規表現に合致しない
    #     elif (
    #         re.match(
    #             r"^[A-Za-z0-9]{1}[A-Za-z0-9_.+-]*@{1}[A-Za-z0-9_.+-]+.[A-Za-z0-9]+$",
    #             input_value[6].strip(),
    #         )
    #         is None
    #     ):
    #         value_error += 1
    #         message += "行番号 " + str(index + 3) + ": 「メールアドレス」を正しくご記入ください。\n"
    #         account_input.cell(row=index + 3, column=7).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # 所属施設①が空欄
    #     if input_value[7] is None:
    #         fill_error += 1
    #         message += "行番号 " + str(index + 3) + ": 「所属施設①」が未記載です。\n"
    #         account_input.cell(row=index + 3, column=8).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # その他の施設を選択かつ所属施設②が空欄
    #     if input_value[7] == "その他の施設" and input_value[8] is None:
    #         fill_error += 1
    #         message += (
    #             "行番号 "
    #             + str(index + 3)
    #             + ": 「所属施設②」が未記載です (国立成育医療研究センター以外のにご所属の場合は必ずご記載ください)。\n"
    #         )
    #         account_input.cell(row=index + 3, column=9).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # 所属部署が空欄
    #     if input_value[9] is None:
    #         fill_error += 1
    #         message += "行番号 " + str(index + 3) + ": 「所属部署」が未記載です。\n"
    #         account_input.cell(row=index + 3, column=10).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # トレーニング受講日が空欄
    #     if input_value[10] is None:
    #         fill_error += 1
    #         message += "行番号 " + str(index + 3) + ": 「トレーニング受講日」が未記載です。\n"
    #         account_input.cell(row=index + 3, column=11).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # トレーニング受講日が日付フォーマットに合致しない
    #     elif not isinstance(input_value[10], datetime.date):
    #         value_error += 1
    #         message += "行番号 " + str(index + 3) + ": 「トレーニング受講日」の記載が指定した日付の記載になっていません。\n"
    #         account_input.cell(row=index + 3, column=11).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # その他の施設を選択かつプロジェクト作成コピー作成の申請権限希望
    #     if input_value[7] == "その他の施設" and input_value[12] == "〇":
    #         value_error += 1
    #         message += (
    #             "行番号 "
    #             + str(index + 2)
    #             + ": 「プロジェクト作成・コピー作成の申請権限」は国立成育医療研究センター職員以外にはに付与できません。\n"
    #         )
    #         account_input.cell(row=index + 3, column=13).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    #     # 新規ユーザの台帳照合
    #     if (
    #         input_value[0] == "持っていない"
    #         and input_value[1] is not None
    #         and input_value[4] is not None
    #         and input_value[5] is not None
    #         and input_value[6] is not None
    #     ):
    #         email_exist = 0
    #         for ledger_value in ledger_values:

    #             # 台帳のメールアドレスと一致 (ただし情報は完全一致しない)
    #             if ledger_value[5] == input_value[6].strip() and (
    #                 ledger_value[2] != input_value[1].strip()
    #                 or ledger_value[3] != input_value[5].strip()
    #                 or ledger_value[4] != input_value[4].strip()
    #             ):
    #                 email_exist += 1

    #         if email_exist:
    #             value_error += 1
    #             message += (
    #                 "行番号 "
    #                 + str(index + 3)
    #                 + ": "
    #                 + input_value[2].strip()
    #                 + " "
    #                 + input_value[3].strip()
    #                 + " さんのメールアドレスは成育REDCapに既に登録されていますが、アカウント情報が登録されているものと異なります。"
    #                 + "既に成育REDCapアカウントをお持ちの場合は、ご利用中のアカウント情報をご記載ください。\n"
    #             )

    #     # 既存ユーザの台帳照合
    #     if (
    #         input_value[0] == "持っている"
    #         and input_value[1] is not None
    #         and input_value[4] is not None
    #         and input_value[5] is not None
    #         and input_value[6] is not None
    #     ):
    #         account_exist = 0
    #         for ledger_value in ledger_values:

    #             # ユーザ名およびメールアドレスが一致 => 既存ユーザ
    #             if (
    #                 ledger_value[2] == input_value[1].strip()
    #                 and ledger_value[5] == input_value[6].strip()
    #             ):
    #                 account_exist += 1

    #         if not account_exist:
    #             value_error += 1
    #             message += (
    #                 "行番号 "
    #                 + str(index + 3)
    #                 + ": "
    #                 + input_value[2].strip()
    #                 + " "
    #                 + input_value[3].strip()
    #                 + " さんの情報が成育REDCapに既に登録されたアカウント情報と異なります。正しく記載をしてください。\n"
    #             )

    # # column-wiseな入力情報の精査
    # responsible_for_contact = [int(value[11] == "〇") for value in input_values]
    # can_apply_for_create = [int(value[12] == "〇") for value in input_values]
    # email_address_array = [
    #     value[6] if value[6] is None else value[6].strip() for value in input_values
    # ]

    # # 連絡担当者の列の〇が0個または2個以上
    # if np.count_nonzero(responsible_for_contact) == 0:
    #     value_error += 1
    #     message += "連絡担当者を1名お選びください。\n"
    #     for i in range(2):
    #         account_input.cell(row=i + 1, column=12).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )
    # elif np.count_nonzero(responsible_for_contact) > 1:
    #     value_error += 1
    #     message += "連絡担当者が複数選択されています。1名にしてください。\n"
    #     for i in range(2):
    #         account_input.cell(row=i + 1, column=12).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    # # 申請権限の付与希望の列の〇が0個または2個以上
    # if np.count_nonzero(can_apply_for_create) == 0:
    #     value_error += 1
    #     message += "プロジェクト作成・コピー作成の申請権限を1名に付与します。希望者を1名お選びください。\n"
    #     for i in range(2):
    #         account_input.cell(row=i + 1, column=13).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )
    # elif np.count_nonzero(can_apply_for_create) > 1:
    #     value_error += 1
    #     message += "プロジェクト作成・コピー作成の申請権限の希望者が複数選択されています。1名にしてください。\n"
    #     for i in range(2):
    #         account_input.cell(row=i + 1, column=13).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    # # Emailアドレスに重複がみられる
    # email_address_set = set(email_address_array)
    # if len(email_address_array) != len(email_address_set):
    #     value_error += 1
    #     message += "記載されたメールアドレスに重複がみられます。複数のユーザで同じメールアドレスを共有することはできません。\n"
    #     for i in range(2):
    #         account_input.cell(row=i + 1, column=7).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )

    # # 記入に不備がある場合
    # if fill_error or value_error:

    #     # 再提出用申請フォーム.xlsxを作成
    #     reject_book_path = (
    #         project_id.strip()
    #         + "_"
    #         + str(datetime.datetime.now().strftime("%Y%m%d"))
    #         + "_再提出用申請フォーム.xlsx"
    #     )
    #     input_book.save(reject_book_path)

    #     # メッセージ加筆
    #     message = (
    #         "ご記入に不備および不明な点があります。\n"
    #         + "添付の「"
    #         + reject_book_path
    #         + "」をご修正のうえ、改めてご提出願います。\n\n"
    #         + "なお、下記に各該当箇所の指摘事項を記載いたします。ご参照ください。\n\n"
    #         + message
    #     )

    #     # 再提出のお願いメッセージ.txtを作成
    #     message_text_path = (
    #         project_id.strip()
    #         + "_"
    #         + str(datetime.datetime.now().strftime("%Y%m%d"))
    #         + "_再提出のお願いメッセージ.txt"
    #     )
    #     with open(message_text_path, "w") as f:
    #         f.write(message)
    #         f.close()

    #     # メッセージ表示しアプリ終了
    #     messagebox.showinfo("お知らせ", "記入に不備があります。\n再提出を依頼してください。\n")
    #     root.quit()
    #     return 0

    # # 記載に不備なし、台帳上書き前にアーカイブ用に保存
    # archive_ledger_path = (
    #     project_id.strip()
    #     + "_"
    #     + str(datetime.datetime.now().strftime("%Y%m%d"))
    #     + "_アーカイブ台帳.xlsx"
    # )
    # ledger_book.save(archive_ledger_path)

    # # 台帳ファイルからプロジェクト台帳を読み込む
    # project_ledger = ledger_book["プロジェクト"]

    # # 台帳ファイルとの照合と事務局修正
    # user_names = [value[2] for value in ledger_values]
    # csv = "Username,First name,Last name,Email address,Institution ID,Sponsor username,Expiration,Comments\n"
    # need_for_inform = 0
    # for index, input_value in enumerate(input_values):

    #     # いくつかの入力値を変数に格納
    #     user_name = input_value[1].strip()
    #     user_name_ja = input_value[2].strip() + " " + input_value[3].strip()
    #     first_name = (
    #         input_value[5]
    #         .strip()
    #         .translate(
    #             str.maketrans({chr(0xFF01 + i): chr(0x0021 + i) for i in range(94)})
    #         )
    #         .capitalize()
    #     )
    #     last_name = (
    #         input_value[4]
    #         .strip()
    #         .translate(
    #             str.maketrans({chr(0xFF01 + i): chr(0x0021 + i) for i in range(94)})
    #         )
    #         .capitalize()
    #     )
    #     if input_value[7] == "その他の施設":
    #         institute = input_value[8].strip()
    #     else:
    #         institute = input_value[7]

    #     # 台帳の値と照合
    #     account_exist = 0
    #     for ledger_value in ledger_values:

    #         # ユーザ名・ローマ字氏名・メールアドレスの一致する既存アカウントが存在するが、成育REDCapアカウントを「持っていない」
    #         if (
    #             input_value[0] == "持っていない"
    #             and ledger_value[2] == user_name
    #             and ledger_value[3] == first_name
    #             and ledger_value[4] == last_name
    #             and ledger_value[5] == input_value[6].strip()
    #         ):
    #             account_exist += 1

    #     # 既存アカウントが存在する
    #     if account_exist:
    #         account_input.cell(row=index + 3, column=1).value = "持っている"
    #         account_input.cell(row=index + 3, column=2).value = user_name
    #         message += (
    #             "行番号 "
    #             + str(index + 3)
    #             + ": "
    #             + user_name_ja
    #             + " さんは既にREDCapに登録されています。既存のアカウント (ユーザー名: "
    #             + user_name
    #             + ") をご利用ください。\n"
    #         )
    #         need_for_inform += 1

    #     # アカウント名のかぶりを解消
    #     if input_value[0] == "持っていない" and len(user_names) > 0:
    #         while (
    #             np.count_nonzero(
    #                 list(
    #                     map(
    #                         lambda x: x == user_name,
    #                         user_names,
    #                     )
    #                 )
    #             )
    #             > 0
    #         ):
    #             user_name = input_value[1].strip + "".join(
    #                 [random.choice(string.digits) for i in range(2)]
    #             )

    #     # アカウント名を事務局修正した場合
    #     if user_name != input_value[1].strip():
    #         account_input.cell(row=index + 3, column=2).value = user_name
    #         account_input.cell(row=index + 3, column=2).fill = PatternFill(
    #             patternType="solid", fgColor="FFFF00"
    #         )
    #         message += (
    #             "行番号 "
    #             + str(index + 3)
    #             + ": "
    #             + user_name_ja
    #             + " さんはの希望されたユーザー名は既に使用されています。"
    #             + user_name
    #             + " を登録ユーザ名としました。\n"
    #         )
    #         need_for_inform += 1

    #     # ユーザ名リストを更新
    #     user_names.append(user_name)

    #     # アカウント台帳に追記
    #     account_ledger.append(
    #         [
    #             datetime.datetime.now(),
    #             project_id.strip(),
    #             user_name,
    #             first_name,
    #             last_name,
    #             input_value[6].strip(),
    #             user_name_ja,
    #             institute,
    #             input_value[9].strip(),
    #             input_value[10],
    #             int(input_value[12] == "〇"),
    #         ]
    #     )

    #     # REDCapインポート用CSVファイルの内容を追加
    #     csv += (
    #         user_name + "," + first_name,
    #         +"," + last_name,
    #         +"," + input_value[6].strip() + ",,,,\n",
    #     )

    #     # 連絡担当者情報抽出
    #     array_of_contact_person = input_values[responsible_for_contact.index(1)]

    #     # プロジェクト台帳に追記
    #     project_ledger.append(
    #         [
    #             datetime.datetime.now(),
    #             project_id.strip(),
    #             project_name.strip(),
    #             array_of_contact_person[2].strip()
    #             + " "
    #             + array_of_contact_person[3].strip(),
    #             array_of_contact_person[6].strip(),
    #         ]
    #     )

    #     # 台帳上書き保存
    #     ledger_book.save(ledger_book_name.get())

    #     # 登録済み申請フォームを作成
    #     account_input.cell(1, 2).value = "確定ユーザー名"
    #     processed_book_path = (
    #         project_id.strip()
    #         + "_"
    #         + str(datetime.datetime.now().strftime("%Y%m%d"))
    #         + "_登録済み申請フォーム.xlsx"
    #     )
    #     input_book.save(processed_book_path)

    #     # REDCapインポート用CSVファイルを作成
    #     csv_path = (
    #         project_id.strip()
    #         + "_"
    #         + str(datetime.datetime.now().strftime("%Y%m%d"))
    #         + "_import.csv"
    #     )
    #     with open(
    #         csv_path,
    #         "w",
    #     ) as f:
    #         f.write(csv)
    #         f.close()

    #     # 事務局修正を行った場合のメッセージ編集
    #     if need_for_inform:
    #         message = (
    #             "ユーザー登録が終了しました。\n"
    #             + "登録内容を記載したエクセルファイルをご確認ください。\n\n"
    #             + "なお、事務局修正を行った箇所があります。\n"
    #             + "エクセルファイルの黄色ハイライトで示した該当箇所および下記のお知らせをご参照ください。\n\n"
    #             + message
    #         )

    #     # 事務局修正を行っていない場合のメッセージ編集
    #     else:
    #         message = "ユーザー登録が終了しました。\n登録内容を記載したエクセルファイルをご確認ください。\n\n" + message

    #     # 登録終了メッセージ.txtを作成
    #     message_text_path = (
    #         project_id.strip()
    #         + "_"
    #         + str(datetime.datetime.now().strftime("%Y%m%d"))
    #         + "_登録終了メッセージ.txt"
    #     )
    #     with open(
    #         message_text_path,
    #         "w",
    #     ) as f:
    #         f.write(message)
    #         f.close()

    # # メッセージ表示しアプリ終了
    # messagebox.showinfo("お知らせ", "処理が終わりました。\n")
    # root.quit()

    # # 例外処理
    # # except:
    # #     messagebox.showerror("エラー", "うまく処理が終了しませんでした。\n")


if __name__ == "__main__":
    root = tk.Tk()
    root.title("REDCapアカウント管理お助けツール")
    root.geometry("600x150")

    # 台帳ファイルのUI
    ledger_book_frame = ttk.Frame(root, padding=10)
    ledger_book_frame.grid()
    ledger_book_set = tk.StringVar()
    ledger_book_set.set("台帳ファイル: ")
    ledger_book_label = ttk.Label(ledger_book_frame, textvariable=ledger_book_set)
    ledger_book_label.grid(row=0, column=0)
    ledger_book_name = tk.StringVar()
    ledger_book_entry = ttk.Entry(
        ledger_book_frame, textvariable=ledger_book_name, width=50
    )
    ledger_book_entry.grid(row=0, column=1)
    ledger_book_button = ttk.Button(
        ledger_book_frame, text="参照", command=select_ledger_book
    )
    ledger_book_button.grid(row=0, column=2)

    # 申請フォームのUI
    input_book_frame = ttk.Frame(root, padding=10)
    input_book_frame.grid()
    input_book_set = tk.StringVar()
    input_book_set.set("申請フォーム: ")
    input_book_label = ttk.Label(input_book_frame, textvariable=input_book_set)
    input_book_label.grid(row=0, column=0)
    input_book_name = tk.StringVar()
    input_book_entry = ttk.Entry(
        input_book_frame, textvariable=input_book_name, width=50
    )
    input_book_entry.grid(row=0, column=1)
    input_book_button = ttk.Button(
        input_book_frame, text="参照", command=select_input_book
    )
    input_book_button.grid(row=0, column=2)

    # ユーザーリストのUI
    user_list_frame = ttk.Frame(root, padding=10)
    user_list_frame.grid()
    user_list_set = tk.StringVar()
    user_list_set.set("ユーザーリスト: ")
    user_list_label = ttk.Label(user_list_frame, textvariable=user_list_set)
    user_list_label.grid(row=0, column=0)
    user_list_name = tk.StringVar()
    user_list_entry = ttk.Entry(user_list_frame, textvariable=user_list_name, width=50)
    user_list_entry.grid(row=0, column=1)
    user_list_button = ttk.Button(user_list_frame, text="参照", command=select_user_list)
    user_list_button.grid(row=0, column=2)

    # 実行ボタンのUI
    run_frame = ttk.Frame(root, padding=10)
    run_frame.grid()
    run_button = ttk.Button(run_frame, text="実行", command=run_process)
    run_button.grid(row=0, column=0)

    root.mainloop()
